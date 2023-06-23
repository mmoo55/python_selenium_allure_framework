import openpyxl
import datetime


class FuncionesExcel:

    def obtener_numero_filas(self, archivo: str, hoja: str):
        """
        Permite obtener el número de filas que se utiliza en el archivo excel.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :return: int
        """
        Wordbook = openpyxl.load_workbook(archivo)
        sheet = Wordbook[hoja]
        return sheet.max_row

    def obtener_numero_columnas(self, archivo: str, hoja: str):
        """
        Permite obtener el número de columnas que se utiliza en el archivo excel.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :return: int
        """
        Wordbook = openpyxl.load_workbook(archivo)
        sheet = Wordbook[hoja]
        return sheet.max_row

    def leer_dato(self, archivo, hoja, row_num: int, column_num: int):
        """
        Permite obtener el dato de una celda.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :param row_num: Número de fila.
        :param column_num: Número de columna.
        :return: str
        """
        Wordbook = openpyxl.load_workbook(archivo)
        sheet = Wordbook[hoja]
        return sheet.cell(row=row_num, column=column_num).value

    def escribir_dato(self, archivo: str, hoja: str, row_num: int, column_num: int, data):
        """
        Permite escribir en el archivo Excel.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :param row_num: Número de fila.
        :param column_num: Número de columna.
        :param data:
        :return:
        """
        Wordbook = openpyxl.load_workbook(archivo)
        sheet = Wordbook[hoja]
        sheet.cell(row=row_num, column=column_num).value = data
        Wordbook.save(archivo)

    def obtener_datos(self, archivo: str, hoja: str):
        """
        Permite obtener los datos de un archivo excel.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :return: list[list[]]
        """
        try:
            cell_ref_i = ""
            column_ref_f = ""
            validate_i = False
            validate_f = False
            lista_datos = []
            excel_file = openpyxl.load_workbook(archivo)
            # Nombre de la hoja
            sheet_name = excel_file[hoja]
            # Número de filas y columnas
            max_rows = sheet_name.max_row
            max_columns = sheet_name.max_column
            # Recorriendo el excel hasta encontrar las referencias
            for i in range(max_rows):
                for j in range(max_columns):
                    cell = sheet_name.cell(i + 1, j + 1)
                    if cell.value == "N":
                        cell_ref_i = cell.offset(row=1, column=0)
                        column_ref_i = cell_ref_i.column_letter
                        validate_i = True
                    if cell.value == "RESULTADO":
                        cell_ref_f = cell.offset(row=1, column=-1)
                        column_ref_f = cell_ref_f.column_letter
                        validate_f = True
            if not validate_i:
                print("No se pudo encontrar el campo 'N'")
                return False, []
            if not validate_f:
                print("No se pudo encontrar el campo 'Resultado'")
                return False, []
            datos = sheet_name[cell_ref_i.coordinate:column_ref_f + str(max_rows)]
            for row in datos:
                datos_fila = []
                cell_number = row[0].value
                if isinstance(cell_number, int):
                    for celda in row[0:len(row)]:
                        campo_valor = celda.value
                        if campo_valor is None:
                            datos_fila.append("")
                        else:
                            if isinstance(campo_valor, datetime.datetime):
                                campo_valor = campo_valor.strftime("%d/%m/%Y")
                            datos_fila.append(campo_valor)
                if datos_fila:
                    lista_datos.append(datos_fila)
            return lista_datos
        except Exception as e:
            print(e)
            print("Error al obtener datos del archivo Excel.")

    def escribir_datos(self, archivo: str, hoja: str, num_fila: int, dato: str):
        """
        Permite escribir texto en una celda de un archivo excel.
        :param dato:Texto que se escribira en el campo RESULTADO
        :param num_fila: Campo N, de tipo número.
        :param archivo: Ruta del archivo excel.
        :param hoja: Nombre de la hoja.
        :return: bool
        """
        try:
            cell_ref = ""
            validate = False
            excel_file = openpyxl.load_workbook(archivo)
            # Nombre de la hoja
            sheet_name = excel_file[hoja]
            # Número de filas y columnas
            max_rows = sheet_name.max_row
            max_columns = sheet_name.max_column
            # Recorriendo el excel hasta encontrar las referencias
            for i in range(max_rows):
                for j in range(max_columns):
                    cell = sheet_name.cell(i + 1, j + 1)
                    if cell.value == "RESULTADO":
                        cell_ref = cell.offset(row=num_fila)
                        validate = True
            if not validate:
                print("No se pudo encontrar el campo 'RESULTADO'")
                return False
            cell_ref.value = dato
            excel_file.save(archivo)
            return True
        except Exception as e:
            print(e)
            print("Error al guardar datos al archivo Excel.")